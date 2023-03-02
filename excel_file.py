#read data from excel file
import openpyxl

workbook = openpyxl.load_workbook('D:\web design and development\python code\user_name.xlsx','r')
worksheet = workbook.active
for row in worksheet.iter_rows(values_only=True):
    print(row)

#write data from sxcel file
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet['A1'] = 'Col1'
worksheet['B1'] = 'Col2'
worksheet['C1'] = 'Col3'
worksheet.append(['Value1', 'Value2', 'Value3'])
worksheet.append(['Value1', 'Value2', 'Value3'])
worksheet.append(['Value1', 'Value2', 'Value3'])
workbook.save('D:\web design and development\python code\Trip-Planner.xlsx')